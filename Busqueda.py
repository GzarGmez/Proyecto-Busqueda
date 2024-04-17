import openpyxl
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import re

def leer_archivo_excel(nombre_archivo, criterio_busqueda):
    tree.delete(*tree.get_children())  # Limpiar tabla anterior
    try:
        libro = openpyxl.load_workbook(nombre_archivo)
        hoja_activa = libro.active
        for fila in hoja_activa.iter_rows():
            if any(criterio_busqueda in str(celda.value) for celda in fila):
                tree.insert("", tk.END, values=[celda.value for celda in fila])
    except FileNotFoundError:
        messagebox.showerror("Error", f"El archivo {nombre_archivo} no fue encontrado.")
    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error: {str(e)}")

def iniciar_sesion():
    usuario = usuario_entry.get()
    contraseña = contraseña_entry.get()

    usuario_valido = validar_usuario(usuario)
    contraseña_valida = validar_contraseña(contraseña)

    if usuario_valido and not contraseña_valida:
        messagebox.showinfo("Información", "Iniciaste Sesión Correctamente :D")
        # Después de iniciar sesión exitosamente, habilitamos la entrada de búsqueda
        dominio_entry.config(state=tk.NORMAL)
        boton_buscar.config(state=tk.NORMAL)
    else:
        mensaje_error = "Inicio de sesión fallido. Problemas con los campos:\n"
        if not usuario_valido:
            mensaje_error += "- El nombre de usuario debe tener al menos 4 caracteres y no contener espacios.\n"
        if contraseña_valida:
            mensaje_error += "- Contraseña:\n"
            for mensaje in contraseña_valida:
                mensaje_error += f"  {mensaje}\n"
        messagebox.showwarning("Advertencia", mensaje_error)

def buscar():
    dominio_a_buscar = dominio_entry.get()
    if dominio_a_buscar:
        nombre_archivo = 'datospersonales.xlsx'
        leer_archivo_excel(nombre_archivo, dominio_a_buscar)
    else:
        messagebox.showwarning("Advertencia", "Ingresa el Dato que deseas buscar")

def validar_usuario(usuario):
    long = r'^\S{4,}$'
    if re.match(long, usuario):
        return True
    else:
        return False

def validar_contraseña(contraseña):
    long = r'^.{7,14}$'
    may = r'[A-Z]'
    minn = r'[a-z]'
    dig = r'\d'
    spchar = r'[^A-Za-z0-9\s]'

    mensajes = []

    if not re.match(long, contraseña):
        mensajes.append("Use entre 8 y 15 caracteres.")
    if not re.search(may, contraseña):
        mensajes.append("Use al menos una letra mayúscula.")
    if not re.search(minn, contraseña):
        mensajes.append("Use al menos una letra minúscula.")
    if not re.search(dig, contraseña):
        mensajes.append("Use al menos un dígito.")
    if not re.search(spchar, contraseña):
        mensajes.append("Use al menos un carácter especial.")
    if re.search(r'\s', contraseña):
        mensajes.append("No use espacios en blanco.")

    return mensajes

def editar_registro():
    seleccion = tree.selection()
    if seleccion:
        item = seleccion[0]
        valores = tree.item(item, 'values')
        # Aquí puedes implementar la lógica para editar el registro utilizando los valores obtenidos
        print("Editar registro:", valores)
        # Ejemplo: Abrir una ventana de edición con los valores actuales para modificarlos
        ventana_edicion = tk.Toplevel()
        ventana_edicion.title("Editar Registro")
        # Crear campos de entrada con los valores actuales para modificarlos
        campos = []
        for i, valor in enumerate(valores):
            ttk.Label(ventana_edicion, text=f"{tree.heading(i)['text']}").grid(row=i, column=0, padx=10, pady=5)
            campo = ttk.Entry(ventana_edicion, textvariable=tk.StringVar(value=valor))
            campo.grid(row=i, column=1, padx=10, pady=5)
            campos.append(campo)

            # Botón para confirmar la edición
            boton_guardar = ttk.Button(ventana_edicion, text="Guardar Cambios",
                                       command=lambda: guardar_cambios(item, campos))
            boton_guardar.grid(row=i + 1, columnspan=2, padx=10, pady=10)



def guardar_cambios(item, campos):
    # Actualizar valores en el Treeview
    nuevos_valores = [campo.get() for campo in campos]
    tree.item(item, values=nuevos_valores)

    # Cargar datos existentes del archivo Excel
    archivo_existente = 'datospersonales_modificado.xlsx'
    libro_existente = openpyxl.load_workbook(archivo_existente)
    hoja_existente = libro_existente.active

    # Crear un nuevo libro de Excel
    nuevo_libro = openpyxl.Workbook()
    nueva_hoja = nuevo_libro.active
    nueva_hoja.append(["FOLIO", "NOMBRE", "CORREO", "TELEFONO"])

    # Agregar todos los datos existentes al nuevo libro de Excel, excepto los que coincidan con los datos editados
    for fila in hoja_existente.iter_rows(min_row=2, max_col=4, values_only=True):
        if list(fila)[:3] not in tree.get_children():
            nueva_hoja.append(fila)

    # Agregar el dato editado al final del nuevo libro de Excel
    nueva_hoja.append(nuevos_valores)

    # Guardar el nuevo archivo Excel
    nuevo_archivo = 'datospersonales_modificado.xlsx'
    nuevo_libro.save(nuevo_archivo)

    messagebox.showinfo("Información", f"Cambios aplicados correctamente. Nuevo archivo guardado como '{nuevo_archivo}'.")

    # Eliminar el registro seleccionado del Treeview
    tree.delete(item)
    ventana_edicion.destroy()

    # Actualizar la tabla con los nuevos datos
    leer_archivo_excel(nuevo_archivo, None)


def eliminar_registro():
    seleccion = tree.selection()
    if seleccion:
        item = seleccion[0]
        valores = tree.item(item, 'values')

        # Cargar datos existentes del archivo Excel
        archivo_existente = 'datospersonales_modificado.xlsx'
        libro_existente = openpyxl.load_workbook(archivo_existente)
        hoja_existente = libro_existente.active

        # Buscar el índice de la fila a eliminar en el archivo Excel
        indice_fila_a_eliminar = None
        for i, fila in enumerate(hoja_existente.iter_rows(min_row=2, max_col=4, values_only=True), start=2):
            if list(fila)[:3] == valores:
                indice_fila_a_eliminar = i
                break

        # Si se encontró el índice de la fila a eliminar, eliminarla del libro de Excel
        if indice_fila_a_eliminar:
            hoja_existente.delete_rows(indice_fila_a_eliminar)

        # Guardar el nuevo archivo Excel
        libro_existente.save(archivo_existente)

        messagebox.showinfo("Información", "Registro eliminado correctamente.")
        tree.delete(item)
    else:
        messagebox.showwarning("Advertencia", "Seleccione un registro para eliminar.")


ventana = tk.Tk()
ventana.title("Inicio de Sesión y Búsqueda")

panel_inicio_sesion = ttk.Frame(ventana)
panel_inicio_sesion.grid(row=0, column=0, padx=10, pady=10)

etiqueta_usuario = ttk.Label(panel_inicio_sesion, text="Nombre de usuario:")
usuario_entry = ttk.Entry(panel_inicio_sesion)
etiqueta_contraseña = ttk.Label(panel_inicio_sesion, text="Contraseña:")
contraseña_entry = ttk.Entry(panel_inicio_sesion, show="*")
boton_iniciar_sesion = ttk.Button(panel_inicio_sesion, text="Iniciar Sesión", command=iniciar_sesion)

etiqueta_usuario.grid(row=0, column=0, padx=10, pady=10)
usuario_entry.grid(row=0, column=1, padx=10, pady=10)
etiqueta_contraseña.grid(row=1, column=0, padx=10, pady=10)
contraseña_entry.grid(row=1, column=1, padx=10, pady=10)
boton_iniciar_sesion.grid(row=2, column=0, columnspan=2, padx=10, pady=10)

panel_busqueda = ttk.Frame(ventana)
panel_busqueda.grid(row=0, column=1, padx=10, pady=10)

etiqueta_busqueda = ttk.Label(panel_busqueda, text="Ingrese el dominio a buscar:")
dominio_entry = ttk.Entry(panel_busqueda, state=tk.DISABLED)
boton_buscar = ttk.Button(panel_busqueda, text="Buscar", command=buscar, state=tk.DISABLED)

etiqueta_busqueda.grid(row=0, column=0, padx=10, pady=10)
dominio_entry.grid(row=0, column=1, padx=10, pady=10)
boton_buscar.grid(row=0, column=2, padx=10, pady=10)

boton_editar = ttk.Button(panel_busqueda, text="Editar", command=editar_registro)
boton_editar.grid(row=1, column=0, padx=10, pady=10)
boton_eliminar = ttk.Button(panel_busqueda, text="Eliminar", command=eliminar_registro)
boton_eliminar.grid(row=1, column=1, padx=10, pady=10)

tree_frame = ttk.Frame(ventana)
tree_frame.grid(row=1, column=0, columnspan=2, padx=10, pady=10)

tree = ttk.Treeview(tree_frame, columns=("FOLIO", "NOMBRE", "CORREO", "TELEFONO"), show="headings")
tree.heading("FOLIO", text="FOLIO")
tree.heading("NOMBRE", text="NOMBRE")
tree.heading("CORREO", text="CORREO")
tree.heading("TELEFONO", text="TELEFONO")

tree.pack()
ventana.mainloop()
